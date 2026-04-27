from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Q
from django.http import HttpResponseForbidden, HttpResponse
from .models import Student, Course, ClassGroup, Attendance, Profile, QRSession, QRScan
from django.contrib.auth.models import User
from django.utils import timezone
import datetime
import uuid
import qrcode
import io
import base64


def get_role(user):
    try:
        return user.profile.role
    except Exception:
        return None


def role_required(*roles):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            if get_role(request.user) not in roles and not request.user.is_superuser:
                return HttpResponseForbidden("Access Denied")
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


def no_student(view_func):
    """Block student role from accessing a view — redirect to their dashboard."""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if get_role(request.user) == 'student':
            return redirect('student_dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


# ── Auth ──────────────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    if request.method == 'POST':
        user = authenticate(request, username=request.POST['username'], password=request.POST['password'])
        if user:
            login(request, user)
            return redirect('home')
        messages.error(request, 'Invalid username or password.')
    return render(request, 'login.html')


def landing(request):
    """Public landing page."""
    return render(request, 'landing.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ── Home / Dashboard ──────────────────────────────────────────────────────────

@login_required
def home(request):
    today = datetime.date.today()
    role = get_role(request.user)
    is_admin = request.user.is_superuser or role == 'admin'

    if is_admin:
        total_students = Student.objects.count()
        total_courses = Course.objects.count()
        present_today = Attendance.objects.filter(date=today, status='Present').count()
        absent_today = Attendance.objects.filter(date=today, status='Absent').count()
        total_today = present_today + absent_today
        pct = round((present_today / total_today * 100), 1) if total_today else 0
        recent = Attendance.objects.select_related('student', 'course').order_by('-date')[:20]
        my_courses = None
    else:
        # Teacher: only their assigned courses
        my_courses = Course.objects.filter(teacher=request.user)
        total_students = Student.objects.filter(courses__in=my_courses).distinct().count()
        total_courses = my_courses.count()
        present_today = Attendance.objects.filter(date=today, status='Present', course__in=my_courses).count()
        absent_today = Attendance.objects.filter(date=today, status='Absent', course__in=my_courses).count()
        total_today = present_today + absent_today
        pct = round((present_today / total_today * 100), 1) if total_today else 0
        recent = Attendance.objects.filter(course__in=my_courses).select_related('student', 'course').order_by('-date')[:20]

    # Student role: redirect to their own dashboard
    role = get_role(request.user)
    if role == 'student':
        try:
            student = request.user.student_profile
            return redirect('student_dashboard')
        except Exception:
            pass

    return render(request, 'home.html', {
        'total_students': total_students,
        'total_courses': total_courses,
        'present_today': present_today,
        'absent_today': absent_today,
        'attendance_pct': pct,
        'recent': recent,
        'today': today,
        'is_admin': is_admin,
        'my_courses': my_courses,
    })


# ── Students ──────────────────────────────────────────────────────────────────

@login_required
@no_student
def student_list(request):
    q = request.GET.get('q', '')
    course_filter = request.GET.get('course', '')
    students_qs = Student.objects.prefetch_related('courses')
    if q:
        students_qs = students_qs.filter(
            Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(student_id__icontains=q)
        )
    if course_filter:
        students_qs = students_qs.filter(courses__id=course_filter)

    # Compute attendance % for each student
    students = []
    for s in students_qs:
        total = Attendance.objects.filter(student=s).count()
        present = Attendance.objects.filter(student=s, status='Present').count()
        pct = round((present / total * 100), 1) if total else 0
        s.att_pct = pct
        s.att_total = total
        students.append(s)

    all_courses = Course.objects.all()
    return render(request, 'students/list.html', {
        'students': students, 'q': q,
        'all_courses': all_courses, 'course_filter': course_filter,
    })


@login_required
@login_required
@role_required('admin')
def student_create(request):
    import re
    courses = Course.objects.all()
    all_classes = ClassGroup.objects.all()
    if request.method == 'POST':
        student_id  = request.POST.get('student_id', '').strip()
        first_name  = request.POST.get('first_name', '').strip()
        last_name   = request.POST.get('last_name', '').strip()
        email       = request.POST.get('email', '').strip()
        login_username = request.POST.get('login_username', '').strip()
        login_password = request.POST.get('login_password', '').strip()

        errors = []

        # First / Last name: letters and spaces only
        if not first_name:
            errors.append('First Name is required.')
        elif not re.fullmatch(r'[A-Za-z ]+', first_name):
            errors.append('First Name must contain letters only.')

        if not last_name:
            errors.append('Last Name is required.')
        elif not re.fullmatch(r'[A-Za-z ]+', last_name):
            errors.append('Last Name must contain letters only.')

        if not email:
            errors.append('Email is required.')
        elif not re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+', email):
            errors.append('Please enter a valid email address.')

        if not student_id:
            errors.append('Student ID is required.')
        elif Student.objects.filter(student_id=student_id).exists():
            errors.append(f'Student ID "{student_id}" already exists.')

        # Login account validation (only if provided)
        if login_username or login_password:
            if not login_username:
                errors.append('Username is required when creating a login.')
            elif not re.match(r'^[A-Za-z]', login_username):
                errors.append('Please insert correct username — must start with a letter (e.g. seya22).')
            elif not re.fullmatch(r'[A-Za-z][A-Za-z0-9]*', login_username):
                errors.append('Username can only contain letters and numbers, starting with a letter.')
            elif User.objects.filter(username=login_username).exists():
                errors.append(f'Username "{login_username}" already exists.')

            if not login_password:
                errors.append('Password is required when creating a login.')
            elif len(login_password) < 6:
                errors.append('Password must be at least 6 characters.')
            elif not re.search(r'[A-Za-z]', login_password):
                errors.append('Password must include at least one letter.')
            elif not re.search(r'\d', login_password):
                errors.append('Password must include at least one number.')
            elif not re.search(r'[!@#$%^&*()\-_=+\[\]{};\'\\:"|,.<>/?`~]', login_password):
                errors.append('Password must include at least one special character (e.g. @, #, !).')

        if errors:
            for e in errors:
                messages.error(request, e)
            return render(request, 'students/form.html', {'courses': courses, 'all_classes': all_classes, 'action': 'Add'})

        # Enforce max 30 students per class section
        class_group_id = request.POST.get('class_group') or None
        section = request.POST.get('section', '').strip()
        if class_group_id:
            current_count = Student.objects.filter(class_group_id=class_group_id).count()
            if current_count >= 30:
                cls_obj = ClassGroup.objects.filter(pk=class_group_id).first()
                messages.error(request, f'Class "{cls_obj}" already has 30 students. Maximum capacity reached.')
                return render(request, 'students/form.html', {'courses': courses, 'all_classes': all_classes, 'action': 'Add'})

        student = Student.objects.create(
            student_id=student_id,
            first_name=first_name,
            last_name=last_name,
            section=section,
            department=request.POST.get('department', '').strip(),
            parent_contact=request.POST.get('parent_contact', '').strip(),
            class_group_id=class_group_id,
        )
        # Save photo if uploaded
        if request.FILES.get('photo'):
            student.photo = request.FILES['photo']
            student.save()
        course_ids = request.POST.getlist('courses')
        if course_ids:
            student.courses.set(course_ids)

        if login_username and login_password:
            login_user = User.objects.create_user(
                username=login_username,
                password=login_password,
                first_name=first_name,
                last_name=last_name,
                email=email,
            )
            Profile.objects.update_or_create(user=login_user, defaults={'role': 'student'})
            student.user = login_user
            student.save()
            messages.success(request, f'Student added with login: {login_username}')
        else:
            messages.success(request, 'Student added successfully.')
        return redirect('student_list')

    return render(request, 'students/form.html', {'courses': courses, 'all_classes': all_classes, 'action': 'Add'})


@login_required
@role_required('admin')
def student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk)
    courses = Course.objects.all()
    all_classes = ClassGroup.objects.all()
    if request.method == 'POST':
        student_id = request.POST.get('student_id', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        if not student_id or not first_name or not last_name:
            messages.error(request, 'Student ID, First Name and Last Name are required.')
            assigned_ids = list(student.courses.values_list('id', flat=True))
            return render(request, 'students/form.html', {'student': student, 'courses': courses, 'all_classes': all_classes, 'assigned_ids': assigned_ids, 'action': 'Edit'})
        if Student.objects.filter(student_id=student_id).exclude(pk=pk).exists():
            messages.error(request, f'Student ID "{student_id}" already used by another student.')
            assigned_ids = list(student.courses.values_list('id', flat=True))
            return render(request, 'students/form.html', {'student': student, 'courses': courses, 'all_classes': all_classes, 'assigned_ids': assigned_ids, 'action': 'Edit'})
        student.student_id = student_id
        student.first_name = first_name
        student.last_name = last_name
        student.section = request.POST.get('section', '').strip()
        student.department = request.POST.get('department', '').strip()
        student.parent_contact = request.POST.get('parent_contact', '').strip()
        new_class_id = request.POST.get('class_group') or None
        # Enforce max 30 per class when changing class assignment
        if new_class_id and str(new_class_id) != str(student.class_group_id):
            current_count = Student.objects.filter(class_group_id=new_class_id).count()
            if current_count >= 30:
                cls_obj = ClassGroup.objects.filter(pk=new_class_id).first()
                messages.error(request, f'Class "{cls_obj}" already has 30 students. Maximum capacity reached.')
                assigned_ids = list(student.courses.values_list('id', flat=True))
                return render(request, 'students/form.html', {
                    'student': student, 'courses': courses, 'all_classes': all_classes,
                    'assigned_ids': assigned_ids, 'action': 'Edit'
                })
        student.class_group_id = new_class_id
        # Save photo if a new one was uploaded
        if request.FILES.get('photo'):
            student.photo = request.FILES['photo']
        student.save()
        course_ids = request.POST.getlist('courses')
        student.courses.set(course_ids)
        messages.success(request, 'Student updated.')
        return redirect('student_list')
    assigned_ids = list(student.courses.values_list('id', flat=True))
    return render(request, 'students/form.html', {'student': student, 'courses': courses, 'all_classes': all_classes, 'assigned_ids': assigned_ids, 'action': 'Edit'})


@login_required
@role_required('admin')
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.delete()
        messages.success(request, 'Student deleted.')
        return redirect('student_list')
    return render(request, 'confirm_delete.html', {'obj': student, 'type': 'Student'})


@login_required
def student_detail(request, pk):
    # Students can only view their own profile
    role = get_role(request.user)
    if role == 'student':
        try:
            own = request.user.student_profile
            if own.pk != pk:
                return redirect('student_dashboard')
        except Exception:
            return redirect('student_dashboard')

    student = get_object_or_404(Student, pk=pk)
    records = Attendance.objects.filter(student=student).select_related('course', 'student__class_group').order_by('-date')
    total = records.count()
    present = records.filter(status='Present').count()
    absent = records.filter(status='Absent').count()
    late = records.filter(status='Late').count()
    pct = round((present / total * 100), 1) if total else 0
    return render(request, 'students/detail.html', {
        'student': student, 'records': records,
        'total': total, 'present': present, 'absent': absent, 'late': late, 'pct': pct
    })


@login_required
def student_dashboard(request):
    """Dedicated dashboard for logged-in students — read-only view of own data."""
    role = get_role(request.user)
    if role not in ('student',) and not request.user.is_superuser:
        return redirect('home')
    try:
        student = request.user.student_profile
    except Exception:
        messages.error(request, 'No student profile linked to your account.')
        return redirect('login')

    records = Attendance.objects.filter(student=student).select_related('course', 'student__class_group').order_by('-date')
    total = records.count()
    present = records.filter(status='Present').count()
    absent = records.filter(status='Absent').count()
    late = records.filter(status='Late').count()
    pct = round((present / total * 100), 1) if total else 0
    return render(request, 'students/dashboard.html', {
        'student': student, 'records': records,
        'total': total, 'present': present, 'absent': absent, 'late': late, 'pct': pct
    })


# ── Courses ───────────────────────────────────────────────────────────────────

@login_required
@no_student
def course_list(request):
    courses = Course.objects.annotate(
        student_count=Count('student')
    ).select_related('teacher', 'class_group').prefetch_related('student_set')
    return render(request, 'courses/list.html', {'courses': courses})


@login_required
@role_required('admin')
def course_create(request):
    teachers = User.objects.filter(profile__role='teacher')
    classes  = ClassGroup.objects.all()
    if request.method == 'POST':
        Course.objects.create(
            course_name=request.POST['course_name'],
            teacher_id=request.POST.get('teacher') or None,
            class_group_id=request.POST.get('class_group') or None,
        )
        messages.success(request, 'Course created.')
        return redirect('course_list')
    return render(request, 'courses/form.html', {'teachers': teachers, 'classes': classes, 'action': 'Add'})


@login_required
@role_required('admin')
def course_edit(request, pk):
    course   = get_object_or_404(Course, pk=pk)
    teachers = User.objects.filter(profile__role='teacher')
    classes  = ClassGroup.objects.all()
    if request.method == 'POST':
        course.course_name   = request.POST['course_name']
        course.teacher_id    = request.POST.get('teacher') or None
        course.class_group_id = request.POST.get('class_group') or None
        course.save()
        messages.success(request, 'Course updated.')
        return redirect('course_list')
    return render(request, 'courses/form.html', {'course': course, 'teachers': teachers, 'classes': classes, 'action': 'Edit'})


@login_required
@role_required('admin')
def course_delete(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.method == 'POST':
        course.delete()
        messages.success(request, 'Course deleted.')
        return redirect('course_list')
    return render(request, 'confirm_delete.html', {'obj': course, 'type': 'Course'})


# ── ClassGroup (Classes) ──────────────────────────────────────────────────────

@login_required
@role_required('admin')
def class_list(request):
    classes = ClassGroup.objects.annotate(
        student_count=Count('students'),
        course_count=Count('courses')
    ).prefetch_related('students', 'courses__teacher').order_by('name', 'section')
    return render(request, 'classes/list.html', {'classes': classes})


@login_required
@role_required('admin')
def class_create(request):
    grade_presets = [f'Grade {i}' for i in range(1, 13)]
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        section = request.POST.get('section', '').strip()
        if not name:
            messages.error(request, 'Class name is required.')
        elif ClassGroup.objects.filter(name=name, section=section).exists():
            messages.error(request, f'Class "{name} {section}" already exists.')
        else:
            ClassGroup.objects.create(name=name, section=section)
            messages.success(request, f'Class "{name}" created.')
            return redirect('class_list')
    return render(request, 'classes/form.html', {'action': 'Add', 'grade_presets': grade_presets})


@login_required
@role_required('admin')
def class_edit(request, pk):
    cls = get_object_or_404(ClassGroup, pk=pk)
    if request.method == 'POST':
        cls.name = request.POST.get('name', '').strip()
        cls.section = request.POST.get('section', '').strip()
        cls.save()
        messages.success(request, 'Class updated.')
        return redirect('class_list')
    return render(request, 'classes/form.html', {'cls': cls, 'action': 'Edit'})


@login_required
@role_required('admin')
def class_delete(request, pk):
    cls = get_object_or_404(ClassGroup, pk=pk)
    if request.method == 'POST':
        cls.delete()
        messages.success(request, 'Class deleted.')
        return redirect('class_list')
    return render(request, 'confirm_delete.html', {'obj': cls, 'type': 'Class'})


# ── Attendance ────────────────────────────────────────────────────────────────

@login_required
@no_student
def attendance_mark(request):
    role = get_role(request.user)
    if role == 'teacher':
        courses = Course.objects.filter(teacher=request.user)
    else:
        courses = Course.objects.all()

    students = []
    selected_course = None
    selected_date = datetime.date.today().isoformat()

    # Collect filter params
    selected_section = request.GET.get('section', '')
    selected_department = request.GET.get('department', '')
    selected_month = request.GET.get('month', '')
    selected_year = request.GET.get('year', '')

    # Build distinct section lists and all classes for dropdowns
    all_sections = Student.objects.exclude(section__isnull=True).exclude(section='').values_list('section', flat=True).distinct().order_by('section')
    all_departments = Student.objects.exclude(department__isnull=True).exclude(department='').values_list('department', flat=True).distinct().order_by('department')
    current_year = datetime.date.today().year
    year_range = list(range(current_year - 4, current_year + 2))

    if request.method == 'GET' and request.GET.get('course'):
        selected_course = get_object_or_404(Course, pk=request.GET['course'])

        # Priority: specific date > month+year > today
        raw_date = request.GET.get('date', '').strip()
        if raw_date:
            # Validate the date string
            try:
                datetime.date.fromisoformat(raw_date)
                selected_date = raw_date
            except ValueError:
                selected_date = datetime.date.today().isoformat()
        elif selected_month and selected_year:
            # No specific date — build from month+year (use today's day if in same month, else 1st)
            try:
                y, m = int(selected_year), int(selected_month)
                today = datetime.date.today()
                if today.year == y and today.month == m:
                    selected_date = today.isoformat()
                else:
                    selected_date = datetime.date(y, m, 1).isoformat()
            except ValueError:
                selected_date = datetime.date.today().isoformat()
        else:
            selected_date = datetime.date.today().isoformat()

        # Get students enrolled in this course (old system) OR in the class linked to this course
        students = Student.objects.filter(courses=selected_course)

        # If no students found via old M2M, try class_group (new system)
        if not students.exists():
            students = Student.objects.all()

        # Apply section filter only if student has a section set
        if selected_section:
            students = students.filter(section__iexact=selected_section)

        # Apply grade/class filter via department field (now stores class name)
        if selected_department:
            students = students.filter(
                Q(department__iexact=selected_department) |
                Q(class_group__name__iexact=selected_department)
            )

        # Pre-fill existing records
        existing = {a.student_id: a.status for a in Attendance.objects.filter(course=selected_course, date=selected_date)}
        for s in students:
            s.existing_status = existing.get(s.id, '')

    if request.method == 'POST':
        # Admin should not mark attendance directly — only teachers do
        _role_check = get_role(request.user)
        if request.user.is_superuser or _role_check == 'admin':
            messages.error(request, 'Admins cannot mark attendance directly. Use Edit / Approve to correct existing records.')
            return redirect('attendance_edit_list')

        course_id = request.POST.get('course')
        date = request.POST.get('date')
        selected_course = get_object_or_404(Course, pk=course_id)
        students_qs = Student.objects.filter(courses=selected_course)
        saved = 0
        for s in students_qs:
            status = request.POST.get(f'status_{s.id}')
            if status in ('Present', 'Absent', 'Late'):
                Attendance.objects.update_or_create(
                    student=s, course=selected_course, date=date,
                    defaults={'status': status}
                )
                saved += 1
        messages.success(request, f'Attendance saved for {saved} students.')
        return redirect('attendance_mark')

    return render(request, 'attendance/mark.html', {
        'courses': courses,
        'students': students,
        'selected_course': selected_course,
        'selected_date': selected_date,
        'selected_section': selected_section,
        'selected_department': selected_department,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'all_sections': all_sections,
        'all_departments': all_departments,
        'all_classes': ClassGroup.objects.all().order_by('name', 'section'),
        'year_range': year_range,
    })


@login_required
@no_student
def attendance_list(request):
    role = get_role(request.user)
    is_admin = request.user.is_superuser or role == 'admin'

    course_id = request.GET.get('course')
    student_id = request.GET.get('student')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if is_admin:
        courses = Course.objects.all()
    else:
        courses = Course.objects.filter(teacher=request.user)

    if not is_admin and not course_id:
        return render(request, 'attendance/list.html', {
            'records': None, 'courses': courses, 'students': [],
            'course_id': '', 'student_id': '', 'date_from': '', 'date_to': '',
            'is_admin': is_admin, 'require_course': True,
            'total_count': 0, 'present_count': 0, 'absent_count': 0,
            'late_count': 0, 'attendance_pct': 0,
        })

    records = Attendance.objects.select_related('student', 'course').order_by('-date')

    if not is_admin:
        records = records.filter(course__in=courses)

    if course_id:
        records = records.filter(course_id=course_id)
    if student_id:
        records = records.filter(student_id=student_id)
    if date_from:
        records = records.filter(date__gte=date_from)
    if date_to:
        records = records.filter(date__lte=date_to)

    if course_id:
        students = Student.objects.filter(courses__id=course_id).distinct()
    elif is_admin:
        students = Student.objects.all()
    else:
        students = Student.objects.filter(courses__in=courses).distinct()

    # Summary counts
    total_count   = records.count()
    present_count = records.filter(status='Present').count()
    absent_count  = records.filter(status='Absent').count()
    late_count    = records.filter(status='Late').count()
    attendance_pct = round(present_count / total_count * 100, 1) if total_count else 0

    return render(request, 'attendance/list.html', {
        'records': records, 'courses': courses, 'students': students,
        'course_id': course_id, 'student_id': student_id,
        'date_from': date_from, 'date_to': date_to,
        'is_admin': is_admin, 'require_course': False,
        'total_count': total_count, 'present_count': present_count,
        'absent_count': absent_count, 'late_count': late_count,
        'attendance_pct': attendance_pct,
    })


# ── Reports ───────────────────────────────────────────────────────────────────

@login_required
@no_student
def report_daily(request):
    date = request.GET.get('date', datetime.date.today().isoformat())
    course_id = request.GET.get('course')
    is_admin = request.user.is_superuser or get_role(request.user) == 'admin'

    if is_admin:
        courses = Course.objects.all()
    else:
        courses = Course.objects.filter(teacher=request.user)

    records = Attendance.objects.filter(date=date).select_related('student', 'course')

    # Teachers only see their own courses
    if not is_admin:
        records = records.filter(course__in=courses)

    if course_id:
        records = records.filter(course_id=course_id)

    present = records.filter(status='Present').count()
    absent = records.filter(status='Absent').count()
    total = present + absent
    pct = round((present / total * 100), 1) if total else 0
    return render(request, 'reports/daily.html', {
        'records': records, 'date': date, 'courses': courses,
        'present': present, 'absent': absent, 'total': total, 'pct': pct,
        'course_id': course_id, 'is_admin': is_admin,
    })


@login_required
@no_student
def report_monthly(request):
    month = request.GET.get('month', datetime.date.today().strftime('%Y-%m'))
    course_id = request.GET.get('course')
    is_admin = request.user.is_superuser or get_role(request.user) == 'admin'

    if is_admin:
        courses = Course.objects.all()
    else:
        courses = Course.objects.filter(teacher=request.user)

    try:
        year, mon = map(int, month.split('-'))
    except Exception:
        year, mon = datetime.date.today().year, datetime.date.today().month

    records = Attendance.objects.filter(date__year=year, date__month=mon).select_related('student', 'course')

    if not is_admin:
        records = records.filter(course__in=courses)

    if course_id:
        records = records.filter(course_id=course_id)

    student_ids = records.values_list('student_id', flat=True).distinct()
    summary = []
    for sid in student_ids:
        s_records = records.filter(student_id=sid)
        total = s_records.count()
        present = s_records.filter(status='Present').count()
        student = s_records.first().student
        summary.append({
            'student': student,
            'total': total,
            'present': present,
            'absent': total - present,
            'pct': round((present / total * 100), 1) if total else 0,
        })

    return render(request, 'reports/monthly.html', {
        'summary': summary, 'month': month, 'courses': courses, 'course_id': course_id,
    })


@login_required
@no_student
def report_datewise(request):
    course_id = request.GET.get('course')
    date = request.GET.get('date', datetime.date.today().isoformat())
    is_admin = request.user.is_superuser or get_role(request.user) == 'admin'

    if is_admin:
        courses = Course.objects.all()
    else:
        courses = Course.objects.filter(teacher=request.user)

    records = []
    selected_course = None
    if course_id:
        selected_course = get_object_or_404(Course, pk=course_id)
        records = Attendance.objects.filter(course_id=course_id, date=date).select_related('student')
        # Teacher can only view their own courses
        if not is_admin and not courses.filter(pk=course_id).exists():
            records = []
            selected_course = None

    return render(request, 'reports/datewise.html', {
        'records': records, 'courses': courses, 'course_id': course_id,
        'date': date, 'selected_course': selected_course,
    })


# ── Teachers ─────────────────────────────────────────────────────────────────

@login_required
@role_required('admin')
def teacher_list(request):
    teachers = User.objects.filter(profile__role='teacher').select_related('profile')
    return render(request, 'teachers/list.html', {'teachers': teachers})


@login_required
@login_required
@role_required('admin')
def teacher_create(request):
    import re
    from django.db import transaction
    courses = Course.objects.all()
    errors = []

    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        password   = request.POST.get('password', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        phone      = request.POST.get('phone', '').strip()
        address    = request.POST.get('address', '').strip()
        course_ids = request.POST.getlist('courses')

        # ── Validation ──────────────────────────────────────────
        # Username: must start with a letter, can contain letters and numbers
        if not username:
            errors.append('Username is required.')
        elif not re.match(r'^[A-Za-z]', username):
            errors.append('Please insert correct username — must start with a letter (e.g. seya22).')
        elif not re.fullmatch(r'[A-Za-z][A-Za-z0-9]*', username):
            errors.append('Username can only contain letters and numbers, starting with a letter.')
        elif User.objects.filter(username=username).exists():
            errors.append(f'Username "{username}" already exists.')

        # Password: min 6 chars, must have letter + digit + special char
        if not password:
            errors.append('Password is required.')
        elif len(password) < 6:
            errors.append('Password must be at least 6 characters.')
        elif not re.search(r'[A-Za-z]', password):
            errors.append('Password must include at least one letter.')
        elif not re.search(r'\d', password):
            errors.append('Password must include at least one number.')
        elif not re.search(r'[!@#$%^&*()_+\-=\[\]{};\'\\:"|,.<>/?`~]', password):
            errors.append('Password must include at least one special character (e.g. @, #, !).')

        # Phone: Ethiopian format — +251XXXXXXXXX or 09XXXXXXXX or 07XXXXXXXX
        if phone:
            et_phone = re.fullmatch(r'(\+2519\d{8}|\+2517\d{8}|09\d{8}|07\d{8})', phone)
            if not et_phone:
                errors.append('Phone must be a valid Ethiopian number (e.g. +251912345678 or 0912345678).')

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            try:
                with transaction.atomic():
                    # Step 1 — Create User
                    user = User.objects.create_user(
                        username=username,
                        password=password,
                        first_name=first_name,
                        last_name=last_name,
                    )
                    # Step 2 — Create Profile with teacher role
                    Profile.objects.update_or_create(
                        user=user,
                        defaults={'role': 'teacher', 'phone': phone, 'address': address},
                    )
                    # Step 3 — Create Teacher record
                    from .models import Teacher as TeacherModel
                    TeacherModel.objects.update_or_create(
                        user=user,
                        defaults={
                            'employee_id':    request.POST.get('employee_id', '').strip() or None,
                            'department':     request.POST.get('department', '').strip(),
                            'specialization': request.POST.get('specialization', '').strip(),
                        }
                    )
                    # Step 4 — Assign Courses
                    if course_ids:
                        Course.objects.filter(id__in=course_ids).update(teacher=user)

                messages.success(request, f'Teacher "{username}" registered successfully.')
                return redirect('teacher_list')
            except Exception as e:
                messages.error(request, f'Error creating teacher: {e}')

    return render(request, 'teachers/form.html', {'courses': courses})


@login_required
@role_required('admin')
def teacher_edit(request, pk):
    teacher = get_object_or_404(User, pk=pk)
    all_courses = Course.objects.all()
    assigned_ids = list(Course.objects.filter(teacher=teacher).values_list('id', flat=True))
    if request.method == 'POST':
        course_ids = request.POST.getlist('courses')
        # Unassign all current courses for this teacher
        Course.objects.filter(teacher=teacher).update(teacher=None)
        # Assign newly selected courses
        if course_ids:
            Course.objects.filter(id__in=course_ids).update(teacher=teacher)
        messages.success(request, f'Courses updated for {teacher.username}.')
        return redirect('teacher_list')
    return render(request, 'teachers/edit.html', {
        'teacher': teacher,
        'courses': all_courses,
        'assigned_ids': assigned_ids,
    })


@login_required
@role_required('admin')
def teacher_delete(request, pk):
    teacher = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        teacher.delete()
        messages.success(request, 'Teacher deleted.')
        return redirect('teacher_list')
    return render(request, 'confirm_delete.html', {'obj': teacher, 'type': 'Teacher'})


@login_required
@no_student
def report_comparative(request):
    students = Student.objects.prefetch_related('courses').all()
    summary = []
    for s in students:
        records = Attendance.objects.filter(student=s)
        total   = records.count()
        present = records.filter(status='Present').count()
        summary.append({
            'student':      s,
            'total':        total,
            'present':      present,
            'absent':       total - present,
            'pct':          round((present / total * 100), 1) if total else 0,
            'course_count': s.courses.count(),
        })
    summary.sort(key=lambda x: x['pct'])
    return render(request, 'reports/comparative.html', {'summary': summary})


# ── Settings ──────────────────────────────────────────────────────────────────

@login_required
@role_required('admin')
def settings_view(request):
    teachers = User.objects.filter(profile__role='teacher').select_related('profile')
    admins   = User.objects.filter(profile__role='admin').select_related('profile')
    students = User.objects.filter(profile__role='student').select_related('profile')
    return render(request, 'settings.html', {
        'teachers': teachers,
        'admins':   admins,
        'students': students,
    })


# ── Student Self-Service Views ────────────────────────────────────────────────

@login_required
def student_attendance_view(request):
    """Student views their own attendance records with filters."""
    role = get_role(request.user)
    if role != 'student':
        return redirect('home')
    try:
        student = request.user.student_profile
    except Exception:
        return redirect('login')

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    subject_filter = request.GET.get('subject', '')

    records = Attendance.objects.filter(student=student).select_related('course').order_by('-date')
    if date_from:
        records = records.filter(date__gte=date_from)
    if date_to:
        records = records.filter(date__lte=date_to)
    if subject_filter:
        records = records.filter(course__course_name__icontains=subject_filter)

    total = records.count()
    present = records.filter(status='Present').count()
    absent = records.filter(status='Absent').count()
    late = records.filter(status='Late').count()
    pct = round((present / total * 100), 1) if total else 0

    courses = Course.objects.filter(student__user=request.user)

    return render(request, 'students/attendance.html', {
        'student': student, 'records': records,
        'total': total, 'present': present, 'absent': absent, 'late': late, 'pct': pct,
        'date_from': date_from, 'date_to': date_to, 'subject_filter': subject_filter,
        'courses': courses,
    })


@login_required
def student_reports_view(request):
    """Student views their own reports and statistics."""
    role = get_role(request.user)
    if role != 'student':
        return redirect('home')
    try:
        student = request.user.student_profile
    except Exception:
        return redirect('login')

    records = Attendance.objects.filter(student=student).select_related('course').order_by('-date')
    total = records.count()
    present = records.filter(status='Present').count()
    absent = records.filter(status='Absent').count()
    late = records.filter(status='Late').count()
    pct = round((present / total * 100), 1) if total else 0

    from django.db.models.functions import TruncMonth
    from django.db.models import Count
    monthly = []
    months_qs = records.annotate(month=TruncMonth('date')).values('month').annotate(
        total=Count('id'),
        present_count=Count('id', filter=Q(status='Present'))
    ).order_by('-month')[:6]
    for m in months_qs:
        t = m['total']
        p = m['present_count']
        monthly.append({
            'month': m['month'].strftime('%B %Y') if m['month'] else '',
            'total': t,
            'present': p,
            'absent': t - p,
            'pct': round((p / t * 100), 1) if t else 0,
        })

    subject_stats = []
    for course in Course.objects.filter(attendance__student=student).distinct():
        r = records.filter(course=course)
        t = r.count()
        p = r.filter(status='Present').count()
        subject_stats.append({
            'name': str(course),
            'total': t, 'present': p,
            'pct': round((p / t * 100), 1) if t else 0,
        })

    return render(request, 'students/reports.html', {
        'student': student, 'total': total, 'present': present,
        'absent': absent, 'late': late, 'pct': pct,
        'monthly': monthly, 'subject_stats': subject_stats,
    })


@login_required
def student_profile_view(request):
    """Student views and updates their own profile."""
    role = get_role(request.user)
    if role != 'student':
        return redirect('home')
    try:
        student = request.user.student_profile
    except Exception:
        return redirect('login')

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'change_password':
            old_pw = request.POST.get('old_password', '')
            new_pw = request.POST.get('new_password', '')
            confirm_pw = request.POST.get('confirm_password', '')
            if not request.user.check_password(old_pw):
                messages.error(request, 'Current password is incorrect.')
            elif new_pw != confirm_pw:
                messages.error(request, 'New passwords do not match.')
            elif len(new_pw) < 6:
                messages.error(request, 'Password must be at least 6 characters.')
            else:
                request.user.set_password(new_pw)
                request.user.save()
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, request.user)
                messages.success(request, 'Password changed successfully.')
        return redirect('student_profile')

    return render(request, 'students/profile.html', {'student': student})


# ── Student Login Management ──────────────────────────────────────────────────

@login_required
@role_required('admin')
def student_create_login(request, pk):
    import re
    student = get_object_or_404(Student, pk=pk)
    if student.user:
        messages.warning(request, f'{student} already has a login account.')
        return redirect('student_detail', pk=pk)

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        errors = []

        if not username:
            errors.append('Username is required.')
        elif not re.match(r'^[A-Za-z]', username):
            errors.append('Please insert correct username — must start with a letter (e.g. seya22).')
        elif not re.fullmatch(r'[A-Za-z][A-Za-z0-9]*', username):
            errors.append('Username can only contain letters and numbers, starting with a letter.')
        elif User.objects.filter(username=username).exists():
            errors.append(f'Username "{username}" already exists.')

        if not password:
            errors.append('Password is required.')
        elif len(password) < 6:
            errors.append('Password must be at least 6 characters.')
        elif not re.search(r'[A-Za-z]', password):
            errors.append('Password must include at least one letter.')
        elif not re.search(r'\d', password):
            errors.append('Password must include at least one number.')
        elif not re.search(r'[!@#$%^&*()\-_=+\[\]{};\'\\:"|,.<>/?`~]', password):
            errors.append('Password must include at least one special character (e.g. @, #, !).')

        if errors:
            for e in errors:
                messages.error(request, e)
            return redirect('student_detail', pk=pk)

        user = User.objects.create_user(
            username=username,
            password=password,
            first_name=student.first_name,
            last_name=student.last_name
        )
        Profile.objects.filter(user=user).update(role='student')
        student.user = user
        student.save()
        messages.success(request, f'Login created for {student}. Username: {username}')
        return redirect('student_detail', pk=pk)

    return redirect('student_detail', pk=pk)


@login_required
@role_required('admin')
def student_manage_courses(request, pk):
    """Admin updates only the enrolled courses for an already-registered student."""
    student = get_object_or_404(Student, pk=pk)
    all_courses = Course.objects.select_related('teacher', 'class_group').all()
    assigned_ids = set(student.courses.values_list('id', flat=True))

    if request.method == 'POST':
        new_ids = request.POST.getlist('courses')
        student.courses.set(new_ids)
        messages.success(request, f'Courses updated for {student.first_name} {student.last_name}.')
        return redirect('student_manage_courses', pk=pk)

    return render(request, 'students/manage_courses.html', {
        'student': student,
        'all_courses': all_courses,
        'assigned_ids': assigned_ids,
    })


# ── Student Role Pages ────────────────────────────────────────────────────────

@login_required
def student_attendance_view(request):
    """Student views their own attendance records with filters."""
    if get_role(request.user) != 'student':
        return redirect('home')
    try:
        student = request.user.student_profile
    except Exception:
        return redirect('login')

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    subject_filter = request.GET.get('subject', '')

    records = Attendance.objects.filter(student=student).select_related('course').order_by('-date')
    if date_from:
        records = records.filter(date__gte=date_from)
    if date_to:
        records = records.filter(date__lte=date_to)
    if subject_filter:
        records = records.filter(course__course_name__icontains=subject_filter)

    total = records.count()
    present = records.filter(status='Present').count()
    absent = records.filter(status='Absent').count()
    late = records.filter(status='Late').count()
    pct = round((present / total * 100), 1) if total else 0

    courses = Course.objects.filter(attendance__student=student).distinct()

    return render(request, 'students/attendance.html', {
        'student': student, 'records': records,
        'total': total, 'present': present, 'absent': absent, 'late': late, 'pct': pct,
        'date_from': date_from, 'date_to': date_to, 'subject_filter': subject_filter,
        'courses': courses,
    })


@login_required
def student_reports_view(request):
    """Student views their own reports and statistics."""
    if get_role(request.user) != 'student':
        return redirect('home')
    try:
        student = request.user.student_profile
    except Exception:
        return redirect('login')

    from django.db.models.functions import TruncMonth

    records = Attendance.objects.filter(student=student).select_related('course').order_by('-date')
    total = records.count()
    present = records.filter(status='Present').count()
    absent = records.filter(status='Absent').count()
    late = records.filter(status='Late').count()
    pct = round((present / total * 100), 1) if total else 0

    # Monthly summary (last 6 months)
    monthly = []
    months_qs = records.annotate(month=TruncMonth('date')).values('month').annotate(
        total=Count('id'),
        present_count=Count('id', filter=Q(status='Present'))
    ).order_by('-month')[:6]
    for m in months_qs:
        t = m['total']
        p = m['present_count']
        monthly.append({
            'month': m['month'].strftime('%B %Y') if m['month'] else '',
            'total': t, 'present': p, 'absent': t - p,
            'pct': round((p / t * 100), 1) if t else 0,
        })

    # Subject-wise breakdown
    subject_stats = []
    for course in Course.objects.filter(attendance__student=student).distinct():
        r = records.filter(course=course)
        t = r.count()
        p = r.filter(status='Present').count()
        subject_stats.append({'name': str(course), 'total': t, 'present': p,
                               'pct': round((p / t * 100), 1) if t else 0})

    return render(request, 'students/reports.html', {
        'student': student, 'total': total, 'present': present,
        'absent': absent, 'late': late, 'pct': pct,
        'monthly': monthly, 'subject_stats': subject_stats,
    })


@login_required
def student_profile_view(request):
    """Student views and updates their own profile."""
    if get_role(request.user) != 'student':
        return redirect('home')
    try:
        student = request.user.student_profile
    except Exception:
        return redirect('login')

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'change_password':
            old_pw = request.POST.get('old_password', '')
            new_pw = request.POST.get('new_password', '')
            confirm_pw = request.POST.get('confirm_password', '')
            if not request.user.check_password(old_pw):
                messages.error(request, 'Current password is incorrect.')
            elif new_pw != confirm_pw:
                messages.error(request, 'New passwords do not match.')
            elif len(new_pw) < 6:
                messages.error(request, 'Password must be at least 6 characters.')
            else:
                request.user.set_password(new_pw)
                request.user.save()
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, request.user)
                messages.success(request, 'Password changed successfully.')
        return redirect('student_profile')

    return render(request, 'students/profile.html', {'student': student})

# ── QR Code Attendance ────────────────────────────────────────────────────────

def _make_qr_image_b64(url):
    """Generate a QR code image and return as base64 string."""
    qr = qrcode.QRCode(version=1, box_size=8, border=4,
                       error_correction=qrcode.constants.ERROR_CORRECT_H)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return base64.b64encode(buf.getvalue()).decode()


@login_required
@no_student
def qr_generate(request):
    """Teacher/Admin generates a QR session."""
    role = get_role(request.user)
    is_admin = request.user.is_superuser or role == 'admin'

    if is_admin:
        courses = Course.objects.all()
        
        classes = ClassGroup.objects.all()
    else:
        courses = Course.objects.filter(teacher=request.user)
        
        classes = ClassGroup.objects.filter(
            Q(courses__teacher=request.user)
        ).distinct()

    session = None
    qr_b64 = None
    scan_url = None

    if request.method == 'POST':
        course_id = request.POST.get('course') or None
        
        class_id = request.POST.get('class_group') or None
        minutes = int(request.POST.get('minutes', 10))
        minutes = max(1, min(minutes, 60))  # clamp 1–60

        expires = timezone.now() + datetime.timedelta(minutes=minutes)
        session = QRSession.objects.create(
            course_id=course_id,
            
            class_group_id=class_id,
            created_by=request.user,
            date=datetime.date.today(),
            expires_at=expires,
        )

        # Build the scan URL
        scan_url = request.build_absolute_uri(f'/attendance/qr/scan/{session.session_id}/')
        qr_b64 = _make_qr_image_b64(scan_url)

    # List recent sessions created by this user
    recent_sessions = QRSession.objects.filter(created_by=request.user).order_by('-created_at')[:10]

    return render(request, 'attendance/qr_generate.html', {
        'courses': courses,
        
        'classes': classes,
        'session': session,
        'qr_b64': qr_b64,
        'scan_url': scan_url,
        'recent_sessions': recent_sessions,
        'is_admin': is_admin,
    })


@login_required
@no_student
def qr_session_toggle(request, session_id):
    """Deactivate/reactivate a QR session."""
    session = get_object_or_404(QRSession, session_id=session_id, created_by=request.user)
    session.is_active = not session.is_active
    session.save()
    status = "activated" if session.is_active else "deactivated"
    messages.success(request, f'Session {status}.')
    return redirect('qr_generate')


@login_required
def qr_scan(request, session_id):
    """Student scans QR → this view marks their attendance."""
    # Only students (or admins testing) can scan
    role = get_role(request.user)

    session = get_object_or_404(QRSession, session_id=session_id)

    # Validate session
    if not session.is_valid():
        return render(request, 'attendance/qr_result.html', {
            'success': False,
            'message': 'This QR code has expired or been deactivated. Ask your teacher to generate a new one.',
        })

    # Get student profile
    try:
        student = request.user.student_profile
    except Exception:
        # Admin/teacher scanning for testing
        if request.user.is_superuser or role in ('admin', 'teacher'):
            return render(request, 'attendance/qr_result.html', {
                'success': False,
                'message': 'QR session is valid. (Staff accounts cannot mark attendance via QR.)',
                'session': session,
            })
        return render(request, 'attendance/qr_result.html', {
            'success': False,
            'message': 'No student profile linked to your account. Contact your admin.',
        })

    # Prevent duplicate scan
    if QRScan.objects.filter(session=session, student=student).exists():
        return render(request, 'attendance/qr_result.html', {
            'success': False,
            'message': 'You have already scanned this QR code. Attendance already recorded.',
            'already_scanned': True,
        })

    # Mark attendance
    att_kwargs = dict(student=student, date=session.date, defaults={'status': 'Present'})
    if session.course_id:
        att_kwargs['course_id'] = session.course_id
    if False:  # subject removed
        pass

    Attendance.objects.update_or_create(**att_kwargs)
    QRScan.objects.create(session=session, student=student)

    return render(request, 'attendance/qr_result.html', {
        'success': True,
        'message': f'Attendance marked as Present for {student.first_name} {student.last_name}!',
        'student': student,
        'session': session,
    })


@login_required
@no_student
def qr_session_detail(request, session_id):
    """Admin/Teacher views who scanned a session."""
    session = get_object_or_404(QRSession, session_id=session_id)
    role = get_role(request.user)
    is_admin = request.user.is_superuser or role == 'admin'
    if not is_admin and session.created_by != request.user:
        return HttpResponseForbidden("Access Denied")

    scans = QRScan.objects.filter(session=session).select_related('student').order_by('scanned_at')
    scan_url = request.build_absolute_uri(f'/attendance/qr/scan/{session.session_id}/')
    qr_b64 = _make_qr_image_b64(scan_url)

    return render(request, 'attendance/qr_session_detail.html', {
        'session': session,
        'scans': scans,
        'qr_b64': qr_b64,
        'scan_url': scan_url,
        'is_valid': session.is_valid(),
    })


# =============================================================================
# EXPORT SYSTEM — role-based data export (CSV / Excel / PDF)
# =============================================================================

import csv
import io as _io
from django.utils.timezone import now as tz_now

# ── helpers ──────────────────────────────────────────────────────────────────

SCHOOL_NAME = "Springfield High School & Elementary School"

def _parse_date(val):
    """Parse a YYYY-MM-DD string; return None on failure."""
    if not val:
        return None
    try:
        return datetime.date.fromisoformat(val)
    except ValueError:
        return None


def _attendance_qs(request, role):
    """
    Return an Attendance queryset scoped to the caller's role.
    Admin  → all records (with optional filters)
    Teacher → only records for subjects/courses they own
    Student → only their own records
    """
    qs = Attendance.objects.select_related('student', 'course', 'recorded_by')

    date_from = _parse_date(request.GET.get('date_from'))
    date_to   = _parse_date(request.GET.get('date_to'))
    class_id  = request.GET.get('class_id')
    search    = request.GET.get('search', '').strip()

    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if class_id:
        qs = qs.filter(student__class_group_id=class_id)
    if search:
        qs = qs.filter(
            Q(student__first_name__icontains=search) |
            Q(student__last_name__icontains=search)
        )

    if role == 'teacher':
        qs = qs.filter(
            
            Q(course__teacher=request.user)
        )
    elif role == 'student':
        try:
            student = request.user.student_profile
            qs = qs.filter(student=student)
        except Exception:
            qs = qs.none()

    return qs.order_by('-date')


def _make_pdf(title, headers, rows, role, user):
    """Generate a PDF using ReportLab and return bytes."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    styles = getSampleStyleSheet()
    accent = colors.HexColor('#3a3fd8')

    title_style = ParagraphStyle('T', parent=styles['Title'],
        fontSize=16, textColor=accent, alignment=TA_CENTER, spaceAfter=4)
    sub_style = ParagraphStyle('S', parent=styles['Normal'],
        fontSize=9, textColor=colors.grey, alignment=TA_CENTER, spaceAfter=12)

    story = [
        Paragraph(SCHOOL_NAME, title_style),
        Paragraph(title, ParagraphStyle('T2', parent=styles['Heading2'],
            fontSize=13, textColor=colors.HexColor('#1e293b'), alignment=TA_CENTER, spaceAfter=2)),
        Paragraph(
            f"Generated: {tz_now().strftime('%Y-%m-%d %H:%M')}  |  "
            f"Role: {role.capitalize()}  |  User: {user.get_full_name() or user.username}  |  "
            f"Total records: {len(rows)}",
            sub_style
        ),
        Spacer(1, 0.3*cm),
    ]

    table_data = [headers] + rows
    col_count  = len(headers)
    col_width  = (landscape(A4)[0] - 3*cm) / col_count

    tbl = Table(table_data, colWidths=[col_width]*col_count, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0), accent),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0), 9),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('FONTSIZE',    (0,1), (-1,-1), 8),
        ('GRID',        (0,0), (-1,-1), 0.4, colors.HexColor('#e2e8f0')),
        ('ROWHEIGHT',   (0,0), (-1,-1), 18),
        ('TOPPADDING',  (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0),(-1,-1), 4),
    ]))
    story.append(tbl)
    doc.build(story)
    buf.seek(0)
    return buf.read()


def _make_excel(title, headers, rows, role, user):
    """Generate an Excel file using openpyxl and return bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    accent_fill = PatternFill('solid', fgColor='3A3FD8')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Meta rows
    ws.append([SCHOOL_NAME])
    ws['A1'].font = Font(bold=True, size=14, color='3A3FD8')
    ws.append([title])
    ws['A2'].font = Font(bold=True, size=12)
    ws.append([
        f"Generated: {tz_now().strftime('%Y-%m-%d %H:%M')}",
        f"Role: {role.capitalize()}",
        f"User: {user.get_full_name() or user.username}",
        f"Total: {len(rows)}",
    ])
    ws.append([])  # blank row

    # Header row
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill   = accent_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Data rows
    for i, row in enumerate(rows):
        ws.append(row)
        fill_color = 'FFFFFF' if i % 2 == 0 else 'F8FAFC'
        for col_idx in range(1, len(row)+1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill      = PatternFill('solid', fgColor=fill_color)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = border

    # Auto-width
    for col_idx in range(1, len(headers)+1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ''))
             for r in range(header_row, ws.max_row+1)),
            default=10
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _make_csv(headers, rows):
    """Generate CSV bytes."""
    buf = _io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode('utf-8-sig')  # BOM for Excel compatibility


# ── Export page (GET) ─────────────────────────────────────────────────────────

@login_required
def export_page(request):
    """Render the export dashboard page."""
    role = get_role(request.user)
    is_admin   = request.user.is_superuser or role == 'admin'
    is_teacher = role == 'teacher'
    is_student = role == 'student'

    classes  = ClassGroup.objects.all() if (is_admin or is_teacher) else ClassGroup.objects.none()
    

    return render(request, 'exports/export_page.html', {
        'role': role,
        'is_admin': is_admin,
        'is_teacher': is_teacher,
        'is_student': is_student,
        'classes': classes,
        
    })


# ── Attendance export (CSV / Excel / PDF) ─────────────────────────────────────

@login_required
def export_attendance(request):
    role     = get_role(request.user)
    is_admin = request.user.is_superuser or role == 'admin'
    fmt      = request.GET.get('format', 'csv').lower()

    # Students can only export their own — PDF only
    if role == 'student' and fmt not in ('pdf',):
        fmt = 'pdf'

    qs = _attendance_qs(request, role)

    headers = ['#', 'Student ID', 'Student Name', 'Class', 'Subject/Course', 'Date', 'Status', 'Remarks']
    rows = []
    for i, a in enumerate(qs, 1):
        rows.append([
            i,
            a.student.student_id or '—',
            str(a.student),
            str(a.student.class_group) if a.student.class_group else '—',
            str(a.subject or a.course or '—'),
            str(a.date),
            a.status,
            a.remarks or '—',
        ])

    title = 'Attendance Report'
    fname = f'attendance_{tz_now().strftime("%Y%m%d_%H%M")}'

    if fmt == 'pdf':
        data = _make_pdf(title, headers, rows, role, request.user)
        resp = HttpResponse(data, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{fname}.pdf"'
    elif fmt == 'excel':
        if role == 'student':
            return HttpResponseForbidden("Students can only export PDF.")
        data = _make_excel(title, headers, rows, role, request.user)
        resp = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{fname}.xlsx"'
    else:  # csv
        if role == 'student':
            return HttpResponseForbidden("Students can only export PDF.")
        data = _make_csv(headers, rows)
        resp = HttpResponse(data, content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="{fname}.csv"'

    return resp


# ── Students export (admin only) ──────────────────────────────────────────────

@login_required
def export_students(request):
    role     = get_role(request.user)
    is_admin = request.user.is_superuser or role == 'admin'
    if not is_admin:
        return HttpResponseForbidden("Access denied.")

    fmt    = request.GET.get('format', 'excel').lower()
    search = request.GET.get('search', '').strip()
    class_id = request.GET.get('class_id')

    qs = Student.objects.select_related('class_group', 'user').all()
    if search:
        qs = qs.filter(Q(first_name__icontains=search) | Q(last_name__icontains=search) | Q(student_id__icontains=search))
    if class_id:
        qs = qs.filter(class_group_id=class_id)

    headers = ['#', 'Student ID', 'First Name', 'Last Name', 'Class', 'Section', 'Department', 'Parent Contact', 'Has Login']
    rows = []
    for i, s in enumerate(qs, 1):
        rows.append([
            i, s.student_id or '—', s.first_name, s.last_name,
            str(s.class_group) if s.class_group else '—',
            s.section or '—', s.department or '—',
            s.parent_contact or '—',
            'Yes' if s.user else 'No',
        ])

    title = 'Students List'
    fname = f'students_{tz_now().strftime("%Y%m%d_%H%M")}'

    if fmt == 'pdf':
        data = _make_pdf(title, headers, rows, role, request.user)
        resp = HttpResponse(data, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{fname}.pdf"'
    elif fmt == 'csv':
        data = _make_csv(headers, rows)
        resp = HttpResponse(data, content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="{fname}.csv"'
    else:
        data = _make_excel(title, headers, rows, role, request.user)
        resp = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{fname}.xlsx"'
    return resp


# ── Teachers export (admin only) ──────────────────────────────────────────────

@login_required
def export_teachers(request):
    role     = get_role(request.user)
    is_admin = request.user.is_superuser or role == 'admin'
    if not is_admin:
        return HttpResponseForbidden("Access denied.")

    fmt = request.GET.get('format', 'excel').lower()

    teachers = User.objects.filter(profile__role='teacher').select_related('profile')
    headers  = ['#', 'Username', 'Full Name', 'Email', 'Phone', 'Department', 'Employee ID', 'Active']
    rows = []
    for i, t in enumerate(teachers, 1):
        prof = getattr(t, 'profile', None)
        tprof = getattr(t, 'teacher_profile', None)
        rows.append([
            i, t.username, t.get_full_name() or '—', t.email or '—',
            prof.phone if prof else '—',
            tprof.department if tprof else '—',
            tprof.employee_id if tprof else '—',
            'Yes' if t.is_active else 'No',
        ])

    title = 'Teachers List'
    fname = f'teachers_{tz_now().strftime("%Y%m%d_%H%M")}'

    if fmt == 'pdf':
        data = _make_pdf(title, headers, rows, role, request.user)
        resp = HttpResponse(data, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{fname}.pdf"'
    elif fmt == 'csv':
        data = _make_csv(headers, rows)
        resp = HttpResponse(data, content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="{fname}.csv"'
    else:
        data = _make_excel(title, headers, rows, role, request.user)
        resp = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{fname}.xlsx"'
    return resp


# ── System report export (admin only) ─────────────────────────────────────────

@login_required
def export_system_report(request):
    role     = get_role(request.user)
    is_admin = request.user.is_superuser or role == 'admin'
    if not is_admin:
        return HttpResponseForbidden("Access denied.")

    fmt = request.GET.get('format', 'excel').lower()

    total_students   = Student.objects.count()
    total_teachers   = User.objects.filter(profile__role='teacher').count()
    total_classes    = ClassGroup.objects.count()
    total_subjects   = Subject.objects.count()
    total_attendance = Attendance.objects.count()
    present_count    = Attendance.objects.filter(status='Present').count()
    absent_count     = Attendance.objects.filter(status='Absent').count()
    late_count       = Attendance.objects.filter(status='Late').count()
    pct = round(present_count / total_attendance * 100, 1) if total_attendance else 0

    headers = ['Metric', 'Value']
    rows = [
        ['Total Students',        total_students],
        ['Total Teachers',        total_teachers],
        ['Total Classes',         total_classes],
        ['Total Subjects',        total_subjects],
        ['Total Attendance Records', total_attendance],
        ['Present Records',       present_count],
        ['Absent Records',        absent_count],
        ['Late Records',          late_count],
        ['Overall Attendance %',  f'{pct}%'],
        ['Report Generated',      tz_now().strftime('%Y-%m-%d %H:%M')],
    ]

    title = 'System Report'
    fname = f'system_report_{tz_now().strftime("%Y%m%d_%H%M")}'

    if fmt == 'pdf':
        data = _make_pdf(title, headers, rows, role, request.user)
        resp = HttpResponse(data, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{fname}.pdf"'
    elif fmt == 'csv':
        data = _make_csv(headers, rows)
        resp = HttpResponse(data, content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="{fname}.csv"'
    else:
        data = _make_excel(title, headers, rows, role, request.user)
        resp = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{fname}.xlsx"'
    return resp


# =============================================================================
# ADMIN: Edit / Approve Attendance Records
# =============================================================================

@login_required
def attendance_edit_list(request):
    """Admin-only: list attendance records with inline edit capability."""
    role = get_role(request.user)
    is_admin = request.user.is_superuser or role == 'admin'
    if not is_admin:
        messages.error(request, 'Access denied. Admins only.')
        return redirect('attendance_list')

    course_id  = request.GET.get('course')
    student_id = request.GET.get('student')
    date_from  = request.GET.get('date_from')
    date_to    = request.GET.get('date_to')
    status_f   = request.GET.get('status')
    approved_f = request.GET.get('approved')  # '0' = pending, '1' = approved, '' = all

    courses  = Course.objects.all()
    students = Student.objects.all()

    records = Attendance.objects.select_related(
        'student', 'course', 'recorded_by', 'approved_by', 'student__class_group'
    ).order_by('-date', 'student__last_name')

    if course_id:
        records = records.filter(course_id=course_id)
    if student_id:
        records = records.filter(student_id=student_id)
    if date_from:
        records = records.filter(date__gte=date_from)
    if date_to:
        records = records.filter(date__lte=date_to)
    if status_f:
        records = records.filter(status=status_f)
    if approved_f == '1':
        records = records.filter(is_approved=True)
    elif approved_f == '0':
        records = records.filter(is_approved=False)

    total_count    = records.count()
    present_count  = records.filter(status='Present').count()
    absent_count   = records.filter(status='Absent').count()
    late_count     = records.filter(status='Late').count()
    approved_count = records.filter(is_approved=True).count()
    pending_count  = records.filter(is_approved=False).count()
    attendance_pct = round(present_count / total_count * 100, 1) if total_count else 0

    return render(request, 'attendance/edit_list.html', {
        'records': records,
        'courses': courses, 'students': students,
        'course_id': course_id, 'student_id': student_id,
        'date_from': date_from, 'date_to': date_to,
        'status_f': status_f, 'approved_f': approved_f,
        'total_count': total_count, 'present_count': present_count,
        'absent_count': absent_count, 'late_count': late_count,
        'approved_count': approved_count, 'pending_count': pending_count,
        'attendance_pct': attendance_pct,
    })


@login_required
def attendance_edit_record(request, pk):
    """Admin-only: edit a single attendance record."""
    role = get_role(request.user)
    is_admin = request.user.is_superuser or role == 'admin'
    if not is_admin:
        return HttpResponseForbidden('Access denied.')

    record = get_object_or_404(
        Attendance.objects.select_related('student', 'course'),
        pk=pk
    )

    if request.method == 'POST':
        new_status  = request.POST.get('status')
        new_remarks = request.POST.get('remarks', '').strip()
        if new_status in ('Present', 'Absent', 'Late'):
            old_status = record.status
            record.status  = new_status
            record.remarks = new_remarks
            record.recorded_by = request.user
            record.save()
            messages.success(
                request,
                f'Record updated: {record.student} — {record.date} '
                f'changed from {old_status} → {new_status}.'
            )
            return redirect(request.POST.get('next', 'attendance_edit_list'))
        else:
            messages.error(request, 'Invalid status value.')

    return render(request, 'attendance/edit_record.html', {
        'record': record,
        'next': request.GET.get('next', 'attendance_edit_list'),
    })


@login_required
def attendance_approve(request, pk):
    """Admin-only: approve a single attendance record (POST with pk>0) or
       bulk-approve selected records (POST with pk=0)."""
    role = get_role(request.user)
    is_admin = request.user.is_superuser or role == 'admin'
    if not is_admin:
        return HttpResponseForbidden('Access denied.')

    if request.method == 'POST':
        if pk == 0:
            ids = request.POST.getlist('record_ids')
            if ids:
                updated = Attendance.objects.filter(pk__in=ids, is_approved=False).update(
                    is_approved=True, approved_by=request.user
                )
                messages.success(request, f'{updated} record(s) approved successfully.')
            else:
                messages.warning(request, 'No records selected for approval.')
        else:
            record = get_object_or_404(Attendance, pk=pk)
            if not record.is_approved:
                record.is_approved = True
                record.approved_by = request.user
                record.save()
                messages.success(
                    request,
                    f'Approved: {record.student} — {record.date} ({record.status})'
                )
            else:
                messages.info(request, 'This record was already approved.')

    return redirect(request.POST.get('next', 'attendance_edit_list'))


@login_required
def student_reset_password(request, pk):
    """Admin-only: reset a student's login password directly from the dashboard."""
    import re
    from django.http import JsonResponse
    role = get_role(request.user)
    is_admin = request.user.is_superuser or role == 'admin'
    if not is_admin:
        return JsonResponse({'ok': False, 'error': 'Access denied.'}, status=403)

    student = get_object_or_404(Student, pk=pk)

    if not student.user:
        return JsonResponse({'ok': False, 'error': f'{student} has no login account to reset.'})

    if request.method == 'POST':
        new_password     = request.POST.get('new_password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()
        errors = []

        if not new_password:
            errors.append('New password is required.')
        elif len(new_password) < 6:
            errors.append('Password must be at least 6 characters.')
        elif not re.search(r'[A-Za-z]', new_password):
            errors.append('Password must include at least one letter.')
        elif not re.search(r'\d', new_password):
            errors.append('Password must include at least one number.')
        elif not re.search(r'[!@#$%^&*()\-_=+\[\]{};\':"\\|,.<>/?`~]', new_password):
            errors.append('Password must include at least one special character.')

        if new_password and confirm_password and new_password != confirm_password:
            errors.append('Passwords do not match.')

        if errors:
            return JsonResponse({'ok': False, 'error': ' '.join(errors)})

        student.user.set_password(new_password)
        student.user.save()
        return JsonResponse({
            'ok': True,
            'message': f'Password reset successfully for {student} (@{student.user.username}).'
        })

    return JsonResponse({'ok': False, 'error': 'Invalid request method.'}, status=405)
